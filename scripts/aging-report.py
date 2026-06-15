"""
Homesavers Scanner - Aging Report (scheduled email)

Emails management a dashboard of how long store-submitted records have been
sitting in PENDING status (today - created_at), so HO back-office turnaround can
be tracked. Three categories are reported separately:

    B = Non-Scans      C = Wrong Prices      D = Wrong Description

- Dashboard (counts per age bucket, per category, worst stores) goes in the
  EMAIL BODY.
- A detailed Excel file (.xlsx) per category is ATTACHED.

Reads pending records from the app's read-only endpoint:
    GET /api/reports/aging   (auth: X-Sync-Secret, same secret as the sync jobs)

Run:
    C:\\Scraping\\homesavers-scanner\\.venv\\Scripts\\python.exe aging-report.py
    python aging-report.py --dry-run          # build HTML, do NOT send
    python aging-report.py --to you@x.com     # send only to one address (test)

Config: aging-report.config.json (see aging-report.config.example.json).
Schedule it with Windows Task Scheduler (see README).
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import json
import os
import smtplib
import sys
from email.message import EmailMessage

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Resolve config + secret next to the script unless overridden.
HERE        = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.environ.get("AGING_REPORT_CONFIG", os.path.join(HERE, "aging-report.config.json"))
SECRET_FILE = r"C:\Homesavers\.sync-secret"
LOG_FILE    = r"C:\Homesavers\logs\aging-report.log"

CATEGORIES = [
    ("B", "Non-Scans"),
    ("C", "Wrong Prices"),
    ("D", "Wrong Description"),
    ("A", "UOM Errors"),
    ("E", "Price Marked Products"),
    ("F", "DRS Errors"),
]
# Task types shown in the "Created in the last 7 days" boxes — the six query
# types plus the two check tasks (Department Check, Price Check).
BOX_TYPES = CATEGORIES + [("J", "Department Check"), ("K", "Price Check")]

# Short labels for the compact boxes.
SHORT = {"A": "UOM Errors", "B": "Non-Scans", "C": "Wrong Prices",
         "D": "Wrong Desc.", "E": "Price Marked", "F": "DRS Errors",
         "J": "Dept Check", "K": "Price Check"}

DEFAULT_BUCKETS = [
    {"label": "0-1 days",  "min": 0,  "max": 1},
    {"label": "2-3 days",  "min": 2,  "max": 3},
    {"label": "4-7 days",  "min": 4,  "max": 7},
    {"label": "8-14 days", "min": 8,  "max": 14},
    {"label": "15+ days",  "min": 15, "max": None},
]


# -- helpers -------------------------------------------------------------------

def log(msg, level="INFO"):
    line = f"{dt.datetime.now():%Y-%m-%d %H:%M:%S} [{level}] {msg}"
    print(line, flush=True)
    try:
        os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass


def load_config():
    if not os.path.exists(CONFIG_PATH):
        log(f"Config not found: {CONFIG_PATH}", "ERROR")
        log("Copy aging-report.config.example.json -> aging-report.config.json and fill it in.", "ERROR")
        sys.exit(1)
    with open(CONFIG_PATH, encoding="utf-8") as f:
        cfg = json.load(f)
    cfg.setdefault("base_url", "https://homesaversscanner.pages.dev")
    cfg.setdefault("aging_buckets", DEFAULT_BUCKETS)
    cfg.setdefault("overdue_days", 3)
    cfg.setdefault("subject_prefix", "Homesavers Aging Report")
    return cfg


def read_secret():
    if not os.path.exists(SECRET_FILE):
        log(f"Secret file not found: {SECRET_FILE}", "ERROR")
        sys.exit(1)
    with open(SECRET_FILE, encoding="utf-8") as f:
        return f.read().strip()


def parse_iso(s):
    """Parse an ISO timestamp (with Z) to an aware datetime."""
    s = (s or "").replace("Z", "+00:00")
    try:
        return dt.datetime.fromisoformat(s)
    except ValueError:
        return None


def bucket_for(age_days, buckets):
    for b in buckets:
        lo = b.get("min", 0)
        hi = b.get("max", None)
        if age_days >= lo and (hi is None or age_days <= hi):
            return b["label"]
    return buckets[-1]["label"] if buckets else "?"


# -- core ----------------------------------------------------------------------

def fetch_records(cfg, secret):
    url = f"{cfg['base_url']}/api/reports/aging"
    resp = requests.get(url, headers={"X-Sync-Secret": secret}, timeout=120)
    resp.raise_for_status()
    data = resp.json()
    now = parse_iso(data.get("now")) or dt.datetime.now(dt.timezone.utc)
    return data.get("records", []), now, data.get("created_last7", {}) or {}


def enrich(records, now, buckets):
    """Add age_days + bucket to each record. Age = calendar days (today - created)."""
    today = now.date()
    out = []
    for r in records:
        created = parse_iso(r.get("created_at"))
        if not created:
            continue
        age = (today - created.date()).days
        if age < 0:
            age = 0
        rr = dict(r)
        rr["age_days"] = age
        rr["bucket"]   = bucket_for(age, buckets)
        rr["created_dt"] = created
        out.append(rr)
    return out


# Homesavers brand greens (rich, dark) used across the email.
G_DEEP   = "#05431F"   # deepest
G_DARK   = "#075E2E"
G_MID    = "#0A7339"
G_RICH   = "#0E9A52"
G_BRIGHT = "#12A156"
HEAD_BG  = "#E8F1EA"
HEAD_TX  = "#064E27"
BORDER   = "#D7DDD5"
TEXT     = "#27322b"
ALERT    = "#B42318"
FONT     = "Segoe UI,Arial,Helvetica,sans-serif"


def build_html(by_cat, cfg, now, created_last7):
    buckets = cfg["aging_buckets"]
    overdue = int(cfg["overdue_days"])
    bucket_labels = [b["label"] for b in buckets]

    total_all   = sum(len(v) for v in by_cat.values())
    overdue_all = sum(1 for v in by_cat.values() for x in v if x["age_days"] > overdue)
    oldest_all  = max((x["age_days"] for v in by_cat.values() for x in v), default=0)
    as_of       = now.astimezone().strftime("%d/%m/%Y %H:%M")
    gen_total   = sum(int(created_last7.get(code, 0)) for code, _ in BOX_TYPES)

    # A KPI box: a rounded cell with a Homesavers-green gradient (solid bgcolor
    # fallback for Outlook desktop, which ignores CSS gradients).
    def box(label, value, g1, g2, value_color="#ffffff"):
        return (
            f'<td width="33%" valign="top" style="padding:6px;">'
            f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'bgcolor="{g1}" style="background:{g1};'
            f'background-image:linear-gradient(135deg,{g1} 0%,{g2} 100%);border-radius:12px;">'
            f'<tr><td style="padding:16px 16px;font-family:{FONT};">'
            f'<div style="font-size:28px;font-weight:700;color:{value_color};line-height:1;">{value}</div>'
            f'<div style="font-size:13px;color:#E7F3EC;margin-top:5px;">{label}</div>'
            f'</td></tr></table></td>'
        )

    # A small "logged yesterday" box — lighter green gradient, dark-green text.
    def sbox(label, value):
        return (
            f'<td valign="top" style="padding:4px;">'
            f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'bgcolor="#D7EBDC" style="background:#D7EBDC;'
            f'background-image:linear-gradient(135deg,#E6F3E9 0%,#C4E2CC 100%);border-radius:10px;">'
            f'<tr><td align="center" style="padding:10px 4px;font-family:{FONT};">'
            f'<div style="font-size:20px;font-weight:700;color:#0A5A2E;line-height:1;">{value}</div>'
            f'<div style="font-size:12px;color:#2C6B42;margin-top:3px;line-height:1.25;">{label}</div>'
            f'</td></tr></table></td>'
        )

    overdue_color = "#FFD24D" if overdue_all > 0 else "#ffffff"
    longest_color = "#FFD24D" if oldest_all > overdue else "#ffffff"

    H = []
    H.append('<!DOCTYPE html><html><head><meta charset="utf-8">'
             '<meta name="viewport" content="width=device-width,initial-scale=1"></head>')
    H.append('<body style="margin:0;padding:0;background:#f1f4f1;">')
    H.append('<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" '
             'style="background:#f1f4f1;"><tr><td align="center" style="padding:18px 12px;">')
    H.append('<table role="presentation" width="600" cellpadding="0" cellspacing="0" border="0" '
             'style="width:600px;max-width:100%;background:#ffffff;border-radius:14px;overflow:hidden;border:1px solid #e2e8e2;">')

    # Header band — green gradient
    H.append(f'<tr><td style="padding:22px 24px;background:{G_DARK};'
             f'background-image:linear-gradient(135deg,{G_RICH} 0%,{G_DEEP} 100%);font-family:{FONT};">'
             f'<div style="font-size:21px;font-weight:700;color:#ffffff;letter-spacing:.2px;">Store Query Status</div>'
             f'<div style="font-size:13px;color:#CDE8D6;margin-top:4px;">As of {as_of}</div></td></tr>')

    # Intro — professional
    H.append(f'<tr><td style="padding:20px 24px 6px;font-family:{FONT};font-size:14px;color:{TEXT};line-height:1.55;">'
             'Good morning,<br><br>'
             'Below is the current status of store queries awaiting action by Support Office &mdash; '
             'how many are still pending, how long they have been waiting, and how many were newly '
             'logged yesterday.</td></tr>')

    # KPI boxes — pending summary, 3 across, gradient stepping deeper
    H.append('<tr><td style="padding:8px 18px 2px;"><table role="presentation" width="100%" '
             'cellpadding="0" cellspacing="0" border="0"><tr>')
    H.append(box("Total pending", total_all, G_BRIGHT, G_MID))
    H.append(box(f"Pending over {overdue} days", overdue_all, G_MID, G_DARK, overdue_color))
    H.append(box("Longest wait (days)", oldest_all, G_DARK, G_DEEP, longest_color))
    H.append('</tr></table></td></tr>')

    # Helpers for the data table — light-green gradient header (not as dark as above)
    def th(t, align="right"):
        return (f'<th align="{align}" style="padding:8px 10px;background:{HEAD_BG};'
                f'background-image:linear-gradient(135deg,#EDF5EE 0%,#D4E8D8 100%);color:{HEAD_TX};'
                f'font-family:{FONT};font-size:12px;font-weight:700;border:1px solid {BORDER};">{t}</th>')
    def td(v, align="right", bold=False, color=TEXT):
        w = "font-weight:700;" if bold else ""
        return (f'<td align="{align}" style="padding:7px 10px;border:1px solid {BORDER};'
                f'font-family:{FONT};font-size:13px;color:{color};{w}">{v}</td>')

    # Pending queries by type
    H.append(f'<tr><td style="padding:16px 24px 2px;font-family:{FONT};font-size:16px;'
             f'font-weight:700;color:#1f2724;">Pending queries by type</td></tr>')
    H.append('<tr><td style="padding:4px 18px 6px;">'
             '<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;">')
    H.append('<tr>' + th("Type", "left") + "".join(th(l) for l in bucket_labels)
             + th("Total") + th("Longest") + '</tr>')
    for code, name in CATEGORIES:
        rows = by_cat.get(code, [])
        counts = {l: 0 for l in bucket_labels}
        for x in rows:
            counts[x["bucket"]] = counts.get(x["bucket"], 0) + 1
        oldest = max((x["age_days"] for x in rows), default=0)
        cells = "".join(td(counts[l] or "") for l in bucket_labels)
        H.append('<tr>' + td(name, "left") + cells + td(len(rows), bold=True) + td(oldest) + '</tr>')
    H.append('</table></td></tr>')

    # Created in the last 7 days — recent activity (all statuses), small boxes by
    # type, including Department Check and Price Check. Placed after the table.
    H.append(f'<tr><td style="padding:16px 24px 0;font-family:{FONT};font-size:16px;'
             f'font-weight:700;color:#1f2724;">Created in the last 7 days'
             f'<span style="font-size:13px;font-weight:400;color:#5b665e;"> &mdash; {gen_total} new across all stores</span></td></tr>')
    H.append('<tr><td style="padding:4px 20px 8px;"><table role="presentation" width="100%" '
             'cellpadding="0" cellspacing="0" border="0"><tr>')
    for code, _ in BOX_TYPES:
        H.append(sbox(SHORT.get(code, code), int(created_last7.get(code, 0))))
    H.append('</tr></table></td></tr>')

    # Footer
    H.append(f'<tr><td style="padding:16px 24px 24px;font-family:{FONT};font-size:13px;'
             f'color:#5b665e;line-height:1.6;border-top:1px solid #eef2ee;">'
             'Detailed line-by-line records are attached as Excel files, one per query type. '
             'Please action the pending queries at your earliest convenience.<br><br>'
             f'Kind regards,<br><strong style="color:{HEAD_TX};">Homesavers Scanner</strong></td></tr>')

    H.append('</table></td></tr></table></body></html>')
    return "\n".join(H)


def build_xlsx(rows):
    """Build an .xlsx workbook (as bytes) of the detailed pending rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pending"
    ws.append(["Store Code", "Store Name", "Product Code", "Description",
               "Quantity", "Submitted (DD/MM/YYYY HH:MM)", "Age (days)"])
    for x in sorted(rows, key=lambda r: -r["age_days"]):
        submitted = x["created_dt"].astimezone().strftime("%d/%m/%Y %H:%M")
        ws.append([x["store_code"], x["store_name"],
                   str(x["product_code"]),          # text keeps leading zeros
                   x["description"], x["quantity"], submitted, x["age_days"]])
    for cell in ws[1]:                              # bold header row
        cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):  # product code = text
        for cell in row:
            cell.number_format = "@"
    ws.freeze_panes = "A2"
    for i, w in enumerate([12, 26, 16, 44, 10, 22, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_email(cfg, html, attachments, recipients, cc=None):
    smtp = cfg["smtp"]
    msg = EmailMessage()
    today = dt.date.today().strftime("%d/%m/%Y")
    msg["Subject"] = f'{cfg["subject_prefix"]} - {today}'
    msg["From"]    = smtp["from"]
    msg["To"]      = ", ".join(recipients)
    if cc:
        msg["Cc"]  = ", ".join(cc)
    msg.set_content("This report needs an HTML-capable email client.")
    msg.add_alternative(html, subtype="html")

    for filename, blob in attachments:
        msg.add_attachment(
            blob,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )

    host, port = smtp["host"], int(smtp.get("port", 587))
    security = (smtp.get("security") or "starttls").lower()
    if security == "ssl":
        with smtplib.SMTP_SSL(host, port, timeout=60) as s:
            if smtp.get("username"):
                s.login(smtp["username"], smtp["password"])
            s.send_message(msg)
    else:
        with smtplib.SMTP(host, port, timeout=60) as s:
            if security == "starttls":
                s.starttls()
            if smtp.get("username"):
                s.login(smtp["username"], smtp["password"])
            s.send_message(msg)


# -- main ----------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="Build the email but do not send; write HTML to disk.")
    ap.add_argument("--to", help="Send only to this address (overrides recipients) - for testing.")
    args = ap.parse_args()

    log("=== Aging report starting ===")
    cfg = load_config()
    secret = read_secret()

    try:
        records, now, created_last7 = fetch_records(cfg, secret)
    except Exception as e:
        log(f"Could not fetch report data: {e}", "ERROR")
        sys.exit(1)
    log(f"Fetched {len(records)} pending records; "
        f"{sum(int(v) for v in created_last7.values())} created in the last 7 days.")

    enriched = enrich(records, now, cfg["aging_buckets"])
    by_cat = {code: [x for x in enriched if x["task_type"] == code] for code, _ in CATEGORIES}

    html = build_html(by_cat, cfg, now, created_last7)

    attachments = []
    for code, name in CATEGORIES:
        rows = by_cat.get(code, [])
        if rows:
            fname = f"{name.replace(' ', '-')}-{dt.date.today():%Y%m%d}.xlsx"
            attachments.append((fname, build_xlsx(rows)))

    if args.dry_run:
        out = os.path.join(HERE, "aging-report-preview.html")
        with open(out, "w", encoding="utf-8") as f:
            f.write(html)
        log(f"DRY RUN - wrote preview to {out}; {len(attachments)} attachment(s) prepared. No email sent.")
        return

    # A --to test send goes to that one address only (no CC); otherwise use the
    # configured To recipients + CC list.
    recipients = [args.to] if args.to else [r for r in cfg.get("recipients", []) if r]
    cc = [] if args.to else [c for c in cfg.get("cc", []) if c]
    if not recipients:
        log("No recipients configured (config.recipients) and no --to given.", "ERROR")
        sys.exit(1)

    try:
        send_email(cfg, html, attachments, recipients, cc)
        log(f"Sent report to {len(recipients)} recipient(s): {', '.join(recipients)}"
            + (f"; cc: {', '.join(cc)}" if cc else ""))
    except Exception as e:
        log(f"Failed to send email: {e}", "ERROR")
        sys.exit(1)

    log("=== Aging report finished OK ===")


if __name__ == "__main__":
    main()
