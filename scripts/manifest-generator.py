"""
Homesavers Scanner - Delivery Manifest generator (local scheduled job)

Picks up each new HSVMAN delivery CSV, looks up Group / Subgroup / Product ID from
the latest ItemMaster (by Short Code = Ud_ImageName), and writes a formatted Excel
manifest that matches the agreed sample:

    Sheets:  Delivery Summary  ->  Master  ->  one sheet per store (alphabetical)

Reads only (no database). All work is local on the back-office PC, like the sync jobs.

Folders:
    INPUT       U:\\...\\Live\\Inbound file\\Delivery Manifest   (HSVMAN*.csv)
    ITEMMASTER  Y:\\...\\ProductMaster\\2026                     (latest ItemMaster*.xlsx)
    OUTPUT      Y:\\Supply Chain\\012 - Manifest\\2026           (Manifest- <Load> - <Trailer>.xlsx)

Run:
    py manifest-generator.py                 # process every NEW HSVMAN (output not already present)
    py manifest-generator.py --dry-run       # build the newest one to a local preview folder, no Y: write
    py manifest-generator.py --file "<path>" # force-process one HSVMAN file
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import glob
import os
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Folders ─────────────────────────────────────────────────────────────────
INPUT_DIR      = r"U:\Finance Dept-Shared Folder\B&M - Homesavers Shared Database\Live\Inbound file\Delivery Manifest"
ITEMMASTER_DIR = r"Y:\Supply Chain & Buying - Shared\Data\VRSDAILYDATADUMP\ProductMaster\2026"
OUTPUT_DIR     = r"Y:\Supply Chain\012 - Manifest\2026"
PREVIEW_DIR    = r"C:\Homesavers\manifest-preview"
LOG_FILE       = r"C:\Homesavers\logs\manifest-generator.log"

# ── Styling (matches the sample workbook) ───────────────────────────────────
FONT_TITLE  = Font(name="Calibri", size=16, bold=True)
FONT_HEAD   = Font(name="Calibri", size=10, bold=True, color="FF000000")
FONT_BOLD   = Font(name="Calibri", size=11, bold=True)
FONT_BODY   = Font(name="Calibri", size=11)
FILL_BLUE   = PatternFill(fill_type="solid", fgColor="FFD9E1F2")  # serial/store/cases/pallets
FILL_AMBER  = PatternFill(fill_type="solid", fgColor="FFFFC000")  # expected arrival
FILL_PEACH  = PatternFill(fill_type="solid", fgColor="FFFCE4D6")  # actual arrival
FILL_YELLOW = PatternFill(fill_type="solid", fgColor="FFFFFF00")  # Master header row
FONT_MHEAD  = Font(name="Calibri", size=11, bold=True)
CENTER      = Alignment(horizontal="center", wrap_text=True)
CENTER_NW   = Alignment(horizontal="center")
LEFT        = Alignment(horizontal="left")
_THIN       = Side(style="thin")
BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
BOTTOM      = Border(bottom=_THIN)


# ── Helpers ─────────────────────────────────────────────────────────────────

def log(msg, level="INFO"):
    line = f"{dt.datetime.now():%Y-%m-%d %H:%M:%S} [{level}] {msg}"
    print(line, flush=True)
    try:
        os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass


def read_text(path):
    """Read a CSV as text, tolerating UTF-8 (with/without BOM) or Windows-1252."""
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    with open(path, "r", encoding="latin-1", errors="replace", newline="") as f:
        return f.read()


def norm_code(v):
    """Normalise a Short Code / Ud_ImageName to a comparable string (digits, no leading zeros)."""
    if v is None:
        return ""
    s = str(v).strip()
    if s == "":
        return ""
    if re.fullmatch(r"\d+", s):
        return str(int(s))          # drop leading zeros so '0402700' == '402700'
    return s.upper()


def to_int(v, default=0):
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return default


def sanitize_sheet_name(name, used):
    clean = re.sub(r"[:\\/?*\[\]]", "_", str(name))[:31]
    base, i = clean, 2
    while clean in used or clean == "":
        suffix = f"_{i}"
        clean = (base[:31 - len(suffix)] + suffix)
        i += 1
    used.add(clean)
    return clean


# ── HSVMAN parsing ──────────────────────────────────────────────────────────

def parse_hsvman(path):
    """Return (header dict, list of ITM dicts). Raw file: quoted, no header, HDR/ITM/TRL rows."""
    rows = list(csv.reader(read_text(path).splitlines()))
    header, items = {}, []
    for r in rows:
        if not r:
            continue
        tag = (r[0] or "").strip().upper()
        if tag == "HDR" and len(r) >= 5:
            header = {
                "dc_name":       r[1].strip(),
                "load_no":       r[2].strip(),
                "trailer_id":    r[3].strip(),
                "trailer_short": r[4].strip(),
            }
        elif tag == "ITM" and len(r) >= 9:
            items.append({
                "pallet_id":   r[1].strip(),
                "store_name":  r[2].strip(),
                "short_code":  r[4].strip(),
                "item_name":   r[5].strip(),
                "total_cases": to_int(r[8]),
            })
        # TRL and anything else: ignore
    # Fall back to the filename for the load number if no HDR was present.
    if not header.get("load_no"):
        m = re.search(r"(LD\d+)", os.path.basename(path), re.IGNORECASE)
        header["load_no"] = m.group(1).upper() if m else "LD_UNKNOWN"
    header.setdefault("trailer_short", "")
    return header, items


def peek_header(path):
    """Cheap read of just the HDR row -> (load_no, trailer_short), to decide what's new."""
    with open(path, "r", encoding="latin-1", newline="") as f:
        first = f.readline()
    r = next(csv.reader([first]), [])
    is_hdr = len(r) > 4 and (r[0] or "").strip().upper() == "HDR"
    load_no = r[2].strip() if is_hdr else ""
    trailer = r[4].strip() if is_hdr else ""
    if not load_no:
        m = re.search(r"(LD\d+)", os.path.basename(path), re.IGNORECASE)
        load_no = m.group(1).upper() if m else "LD_UNKNOWN"
    return load_no, trailer


# ── ItemMaster lookup ───────────────────────────────────────────────────────

def load_itemmaster_lookup(path):
    """Build { norm(Ud_ImageName): (Product ID/EAN_Barcode, Group/ItemGroup, Subgroup/ItemSubGrp_Id) }."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = next(it)
    want = ("Ud_ImageName", "EAN_Barcode", "ItemGroup", "ItemSubGrp_Id")
    idx = {h: i for i, h in enumerate(header) if h in want}
    missing = [w for w in want if w not in idx]
    if missing:
        wb.close()
        raise ValueError(f"ItemMaster missing columns {missing}. Found: {list(header)[:40]}")
    lookup = {}
    iU, iE, iG, iS = idx["Ud_ImageName"], idx["EAN_Barcode"], idx["ItemGroup"], idx["ItemSubGrp_Id"]
    for r in it:
        key = norm_code(r[iU])
        if not key or key in lookup:
            continue
        lookup[key] = (
            (str(r[iE]).strip() if r[iE] is not None else ""),
            (str(r[iG]).strip() if r[iG] is not None else ""),
            (str(r[iS]).strip() if r[iS] is not None else ""),
        )
    wb.close()
    return lookup


def latest(pattern_dir, pattern):
    files = glob.glob(os.path.join(pattern_dir, pattern))
    return max(files, key=os.path.getmtime) if files else None


# ── Workbook building ───────────────────────────────────────────────────────

MASTER_HEADERS = ["Seq. No", "Load No", "Trailer ID", "Pallet ID", "Store Name", "Product ID",
                  "Short Code", "Item Name", "Total Cases", "Group", "Subgroup",
                  "Additional Comments", "Product ID"]
STORE_HEADERS  = ["Load No", "Trailer ID", "Pallet ID", "Store Name", "Product ID", "Short Code",
                  "Item Name", "Total Cases", "Group", "Subgroup", "firl1"]
MASTER_WIDTHS  = [8, 12.9, 13.7, 14.7, 22.7, 14.7, 15.7, 30.3, 16, 24.9, 28.7, 36.7, 14.7]
STORE_WIDTHS   = [12, 11, 16, 22, 14.7, 12, 30, 12, 16, 28, 8]


def enrich(items, lookup):
    """Attach Product ID / Group / Subgroup to each ITM row from the ItemMaster lookup.
    Unmatched short codes are flagged and shown as 'Not in Item Master' in the output."""
    out = []
    for it in items:
        hit = lookup.get(norm_code(it["short_code"]))
        row = dict(it)
        row["matched"] = hit is not None
        row["product_id"], row["group"], row["subgroup"] = hit if hit else ("Not in Item Master", "", "")
        row["short_int"] = to_int(it["short_code"], None)
        out.append(row)
    return out


def build_workbook(header, rows):
    load_no = header["load_no"]
    trailer = header.get("trailer_short") or ""

    # Store order = first appearance in the data (summary + pallet list);
    # per-store sheets are alphabetical (as in the sample).
    order, seen = [], set()
    for r in rows:
        s = r["store_name"]
        if s not in seen:
            seen.add(s); order.append(s)
    cases = {s: 0 for s in order}
    pallets = {s: [] for s in order}
    rowcount = {s: 0 for s in order}
    for r in rows:
        s = r["store_name"]
        cases[s] += r["total_cases"]
        rowcount[s] += 1
        if r["pallet_id"] and r["pallet_id"] not in pallets[s]:
            pallets[s].append(r["pallet_id"])

    ref_store = max(rowcount, key=rowcount.get) if rowcount else None   # most rows -> firl1 False

    wb = Workbook()
    _build_summary(wb.active, load_no, trailer, order, cases, pallets)
    _build_master(wb.create_sheet("Master"), load_no, trailer, rows, ref_store)
    used = {"Delivery Summary", "Master"}
    for store in sorted(order):
        ws = wb.create_sheet(sanitize_sheet_name(store, used))
        _build_store(ws, load_no, trailer, [r for r in rows if r["store_name"] == store])
    return wb, load_no, trailer


def _build_summary(ws, load_no, trailer, order, cases, pallets):
    ws.title = "Delivery Summary"
    ws.sheet_view.showGridLines = False
    ws["A2"] = f"Delivery Summary - {load_no} - {trailer}"
    ws["A2"].font = FONT_TITLE
    ws.row_dimensions[2].height = 28

    # Left table: Store Name | Total Cases.
    ws["A5"] = "Store Name"; ws["B5"] = "Total Cases."
    ws["A5"].font = ws["B5"].font = FONT_BODY
    r = 6
    for s in order:
        ws.cell(r, 1, s).font = FONT_BODY; ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, cases[s]).font = FONT_BODY
        r += 1
    ws.cell(r, 1, "Grand Total").font = FONT_BODY
    ws.cell(r, 2, sum(cases.values())).font = FONT_BODY

    # Right table: heading + a bordered grid whose colour blocks run DOWN through
    # the data rows (Expected=amber, Actual=peach, Status=white) — matches sample.
    ws["D4"] = "Delivery Summary by Store"; ws["D4"].font = FONT_TITLE
    rt_heads = ["Serial no.", "Store Name", "Total Cases", "Total Pallets",
                "Expected Day of Arrival at Dublin Port", "Expected Time of Arrival at Dublin Port",
                "Actual Day of Arrival to stores", "Actual Time of Arrival to stores", "Delivery Status"]
    hdr_fill  = {4: FILL_BLUE, 5: FILL_BLUE, 6: FILL_BLUE, 7: FILL_BLUE,
                 8: FILL_AMBER, 9: FILL_AMBER, 10: FILL_PEACH, 11: FILL_PEACH, 12: None}
    data_fill = {8: FILL_AMBER, 9: FILL_AMBER, 10: FILL_PEACH, 11: FILL_PEACH}
    for col in range(4, 13):
        c = ws.cell(6, col, rt_heads[col - 4])
        c.font = FONT_HEAD; c.alignment = CENTER; c.border = BORDER
        if hdr_fill[col]:
            c.fill = hdr_fill[col]
    ws.row_dimensions[6].height = 40
    rr = 7
    for i, s in enumerate(order, start=1):
        ws.cell(rr, 4, i); ws.cell(rr, 5, s); ws.cell(rr, 6, cases[s]); ws.cell(rr, 7, len(pallets[s]))
        for col in range(4, 13):
            c = ws.cell(rr, col)
            c.font = FONT_BODY; c.alignment = CENTER_NW; c.border = BORDER
            if col in data_fill:
                c.fill = data_fill[col]
        rr += 1
    ws.cell(rr, 5, "Grand Total"); ws.cell(rr, 6, sum(cases.values()))
    ws.cell(rr, 7, sum(len(p) for p in pallets.values()))
    for col in (4, 5, 6, 7):
        ws.cell(rr, col).font = FONT_BOLD; ws.cell(rr, col).alignment = CENTER_NW; ws.cell(rr, col).border = BOTTOM

    # Pallet list: 3 blank rows after the right table, then a heading + per-store blocks.
    pr = rr + 4
    ws.cell(pr, 1, "Store Name").font = FONT_BODY
    pr += 1
    for s in order:
        ws.cell(pr, 1, s).font = FONT_BOLD; ws.cell(pr, 1).alignment = LEFT; pr += 1
        for pid in pallets[s]:
            ws.cell(pr, 1, pid).font = FONT_BODY; ws.cell(pr, 1).alignment = LEFT; pr += 1
    ws.cell(pr, 1, "Grand Total").font = FONT_BOLD
    ws.cell(pr, 2, sum(len(p) for p in pallets.values())).font = FONT_BOLD

    for col, w in {"A": 25.4, "B": 12, "E": 20.6, "F": 9.6, "G": 10.4,
                   "H": 31, "I": 32, "J": 24.6, "K": 25.7, "L": 16.7}.items():
        ws.column_dimensions[col].width = w


def _build_master(ws, load_no, trailer, rows, ref_store):
    ws.sheet_view.showGridLines = False
    ws["A2"] = f"Manifest - {load_no} - {trailer}"; ws["A2"].font = FONT_TITLE
    for i, h in enumerate(MASTER_HEADERS):
        c = ws.cell(5, i + 1, h)
        c.font = FONT_MHEAD; c.fill = FILL_YELLOW; c.border = BORDER
    r = 6
    for seq, row in enumerate(rows, start=1):
        flag = (row["store_name"] != ref_store) if ref_store is not None else False
        vals = [seq, load_no, trailer, row["pallet_id"], row["store_name"], row["product_id"],
                row["short_int"], row["item_name"], row["total_cases"], row["group"],
                row["subgroup"], flag, row["product_id"]]
        for i, v in enumerate(vals):
            c = ws.cell(r, i + 1, v)
            c.font = FONT_BODY; c.border = BORDER
        r += 1
    for i, w in enumerate(MASTER_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def _build_store(ws, load_no, trailer, rows):
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(STORE_HEADERS):
        ws.cell(1, i + 1, h).font = FONT_BODY
    r = 2
    for row in rows:
        vals = [load_no, trailer, row["pallet_id"], row["store_name"], row["product_id"],
                row["short_int"], row["item_name"], row["total_cases"], row["group"],
                row["subgroup"], False]
        for i, v in enumerate(vals):
            ws.cell(r, i + 1, v).font = FONT_BODY
        r += 1
    for i, w in enumerate(STORE_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:K{ws.max_row}"   # filter dropdowns like the original


# ── Rejection report ────────────────────────────────────────────────────────

def rejection_name(load_no, trailer):
    return f"Manifest- {load_no} - {trailer} - Rejections.xlsx"


def write_rejections(rejects, load_no, trailer, out_dir):
    """One row per unmatched Short Code: Short Code | Description | Reason."""
    seen, uniq = set(), []
    for r in rejects:
        key = norm_code(r["short_code"])
        if key in seen:
            continue
        seen.add(key); uniq.append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rejections"
    ws.sheet_view.showGridLines = False
    ws["A2"] = f"Rejections - {load_no} - {trailer}"; ws["A2"].font = FONT_TITLE
    headers = ["Short Code", "Description", "Reason for Rejection"]
    for i, h in enumerate(headers):
        c = ws.cell(4, i + 1, h)
        c.font = FONT_MHEAD; c.fill = FILL_YELLOW; c.border = BORDER
    r = 5
    for row in uniq:
        vals = [row["short_int"] if row["short_int"] is not None else row["short_code"],
                row["item_name"],
                "This Short Code is not in the VRS Item Master"]
        for i, v in enumerate(vals):
            c = ws.cell(r, i + 1, v)
            c.font = FONT_BODY; c.border = BORDER
        r += 1
    for col, w in zip("ABC", (14, 36, 44)):
        ws.column_dimensions[col].width = w
    out_path = os.path.join(out_dir, rejection_name(load_no, trailer))
    wb.save(out_path)
    return out_path, len(uniq)


# ── Process one file ────────────────────────────────────────────────────────

def output_name(load_no, trailer):
    return f"Manifest- {load_no} - {trailer}.xlsx"


def process_file(path, lookup, out_dir):
    header, items = parse_hsvman(path)
    if not items:
        log(f"SKIP (no ITM rows): {os.path.basename(path)}", "WARN")
        return None
    rows = enrich(items, lookup)
    wb, load_no, trailer = build_workbook(header, rows)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, output_name(load_no, trailer))
    wb.save(out_path)
    matched = sum(1 for r in rows if r["matched"])
    rate = 100.0 * matched / len(rows) if rows else 0.0
    stores = len({r["store_name"] for r in rows})
    log(f"WROTE {os.path.basename(out_path)} | rows={len(rows)} stores={stores} match={rate:.1f}%")
    rejects = [r for r in rows if not r["matched"]]
    if rejects:
        rej_path, n = write_rejections(rejects, load_no, trailer, out_dir)
        log(f"WROTE {os.path.basename(rej_path)} | {n} short code(s) not in the VRS Item Master", "WARN")
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="Write to a local preview folder instead of Y:.")
    ap.add_argument("--file", help="Force-process one HSVMAN file path.")
    ap.add_argument("--days", type=int, default=2, help="Only consider HSVMAN files from the last N days (default 2 = yesterday+today).")
    ap.add_argument("--all", action="store_true", help="Ignore the day window (process every HSVMAN; full backfill).")
    args = ap.parse_args()

    log("=== Manifest generator starting ===")
    im_path = latest(ITEMMASTER_DIR, "ItemMaster*.xlsx")
    if not im_path:
        log(f"No ItemMaster found in {ITEMMASTER_DIR}", "ERROR"); sys.exit(1)

    def get_lookup():
        log(f"ItemMaster: {os.path.basename(im_path)}")
        try:
            lk = load_itemmaster_lookup(im_path)
        except Exception as e:
            log(f"Could not load ItemMaster: {e}", "ERROR"); sys.exit(1)
        log(f"ItemMaster lookup rows: {len(lk):,}")
        return lk

    out_dir = PREVIEW_DIR if args.dry_run else OUTPUT_DIR

    if args.file:
        process_file(args.file, get_lookup(), out_dir)
        log("=== Manifest generator finished ==="); return

    files = glob.glob(os.path.join(INPUT_DIR, "HSVMAN*.csv")) + glob.glob(os.path.join(INPUT_DIR, "HSVMAN*.CSV"))
    files = sorted(set(files), key=os.path.getmtime)
    if not args.all:
        cutoff = dt.date.today() - dt.timedelta(days=max(0, args.days - 1))
        files = [f for f in files if dt.date.fromtimestamp(os.path.getmtime(f)) >= cutoff]
    # Decide the work list BEFORE loading the 33MB ItemMaster (so idle runs are instant).
    # Live runs skip loads already generated; dry-runs always (re)build to preview.
    todo = []
    for f in files:
        ln, tr = peek_header(f)
        if args.dry_run or not os.path.exists(os.path.join(out_dir, output_name(ln, tr))):
            todo.append(f)
    if not todo:
        log("No HSVMAN files to process."); log("=== Manifest generator finished ==="); return
    log(f"{len(todo)} file(s) to process -> {out_dir}")
    lookup = get_lookup()
    done = 0
    for f in todo:
        try:
            if process_file(f, lookup, out_dir):
                done += 1
        except Exception as e:
            log(f"FAILED {os.path.basename(f)}: {e}", "ERROR")
    log(f"Processed {done} manifest(s).")
    log("=== Manifest generator finished ===")


if __name__ == "__main__":
    main()
